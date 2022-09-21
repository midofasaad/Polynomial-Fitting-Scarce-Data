"""
Mahmoud Saad 
EA-356
18.01.2022

"""
import os
import openpyxl
import numpy as np



path = os.path.dirname(__file__)

data_sheet_path = path +'/data/measurements/'
file_name = 'Semikron_SiC_1200V_5xST_025deg_Forward.xlsx'
excel_sheet_name=['Tabelle1','Tabelle2','Tabelle3','Tabelle4','Tabelle5']


# Id Vds t
wb_col=[[49,51],[49,51], [49,51],[49,51],[49,51]]
wb_row=[5,62]
#wb_row=[5,500]

#%%


wb_Rds_on = openpyxl.load_workbook(data_sheet_path + file_name)

data=[]
for sheets,col in zip(excel_sheet_name,wb_col):
    IU_data=np.zeros((wb_row[1]-wb_row[0],len(col)))
    ws_Rds_on = wb_Rds_on[sheets]
    for cn,curr_col in enumerate(col):
        for idx in range(wb_row[0],wb_row[1]): # get time current and voltage
            # MAP_data[idx,curr_col] = wb_Rds_on.cell_value(idx, curr_col) # Read the data in the current cell
            IU_data[idx-wb_row[0],cn] = ws_Rds_on.cell(idx+1,curr_col+1).value # Read the data in the current cell
    data.append(IU_data)
#%%

wb_col=[[49,51],[49,51], [49,51],[49,51],[49,51]]
wb_row=[5,56]
wb_Rds_on = openpyxl.load_workbook(data_sheet_path + file_name)

data_2=[]
for sheets,col in zip(excel_sheet_name,wb_col):
    IU_data=np.zeros((wb_row[1]-wb_row[0],len(col)))
    ws_Rds_on = wb_Rds_on[sheets]
    for cn,curr_col in enumerate(col):
        for idx in range(wb_row[0],wb_row[1]): # get time current and voltage
            # MAP_data[idx,curr_col] = wb_Rds_on.cell_value(idx, curr_col) # Read the data in the current cell
            IU_data[idx-wb_row[0],cn] = ws_Rds_on.cell(idx+1,curr_col+1).value # Read the data in the current cell
    data_2.append(IU_data)
    
#%%



data_dic={ 
    'Id_25_degC_1200V'  :data[0][:,0],
    'Vds_25_degC_1200V' :data[0][:,1],
    
    'Id_85_degC_1200V' :data[1][:,0],
    'Vds_85_degC_1200V':data[1][:,1], 
    
    'Id_125_degC_1200V' :data[2][:,0],
    'Vds_125_degC_1200V':data[2][:,1],
    
    'Id_150_degC_1200V' :data[3][:,0],
    'Vds_150_degC_1200V':data[3][:,1],
    
    'Id_175_degC_1200V' :data_2[4][:,0],
    'Vds_175_degC_1200V':data_2[4][:,1] 
    
    
    }


np.savez('VI_curve_measurement.npz', **data_dic)